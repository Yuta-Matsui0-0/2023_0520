import os
import docx
import openpyxl
from datetime import datetime

# 入力データと出力データのパス
input_folder_path = "./Input"
output_folder_path = "./Output"

# 見出しと追加する文言のデフォルト値
headers = ["ABC", "DEF", "GHI", "JKL"]
default_messages = ["", "", "", ""]

# Excelファイル出力用のリスト
output_data = []

# 追加する文言を取得する関数
def get_messages():
    messages = []
    for i in range(len(headers)):
        message = input(f"{headers[i]}直下に追加する文言を設定して下さい。（現在の値：{default_messages[i]}）：")
        if message:
            messages.append(message)
        else:
            messages.append(default_messages[i])
    return messages

# 複数のWordファイルを読み込んで、指定された見出し直下に文言を追加する関数
def process_word_files():
    for filename in os.listdir(input_folder_path):
        if filename.endswith(".docx"):
            # Wordファイルを読み込む
            doc = docx.Document(os.path.join(input_folder_path, filename))
            # Excelファイル出力用のデータを作成するために、初期値を設定する
            row_data = [filename]
            for i in range(len(headers)):
                row_data.append("-")
            # 指定された見出し直下に文言を追加する
            for para in doc.paragraphs:
                for i in range(len(headers)):
                    if para.text == headers[i]:
                        messages = get_messages()
                        for message in messages:
                            para.insert_paragraph_before(message)
                        row_data[i+1] = "v"
            # 出力データを追加する
            output_data.append(row_data)

# Excelファイルを出力する関数
def output_excel_file():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Result"
    # ヘッダーを出力する
    sheet.cell(row=1, column=1).value = "File name"
    for i in range(len(headers)):
        sheet.cell(row=1, column=i+2).value = headers[i]
    # データを出力する
    for i in range(len(output_data)):
        for j in range(len(output_data[i])):
            sheet.cell(row=i+2, column=j+1).value = output_data[i][j]
    # ファイルを保存する
    output_file_name = "InsertResult_" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"
    wb.save(os.path.join(output_folder_path, output_file_name))

if __name__ == '__main__':
    # 追加する文言のデフォルト値を取得する
    default_messages = ["input_default_message" for i in range(len(headers))]

    # 複数のWordファイルを読み込んで、指定された見出し直下に文言を追加する
    process_word_files()

    # Excelファイルを出力する
    output_excel_file()

    # 処理が終了した旨を表示する
    print("処理が終了しました。")
