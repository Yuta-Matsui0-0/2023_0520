import os
import pandas as pd
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 検索ディレクトリと出力ディレクトリのパス
input_dir = 'path_to_your_input_directory'
output_dir = 'path_to_your_output_directory'

# 新しいワークブックを作成
wb = Workbook()

# ヘッダを作成
ws = wb.active
ws.append(["File name", "A_sample", "", "B_sample", ""])
ws.append(["", "シートの有無", "情報の有無", "シートの有無", "情報の有無"])

# 入力フォルダ内の全てのExcelファイルを走査
for file_name in os.listdir(input_dir):
    if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
        file_path = os.path.join(input_dir, file_name)

        # openpyxlを使用してワークブックを開く
        wb_in = load_workbook(filename = file_path, read_only=True)

        # A_sampleシートの確認
        if "A_sample" in wb_in.sheetnames:
            a_sheet_exists = "v"
            a_data_exists = "v" if wb_in["A_sample"]["C3"].value else "-"
        else:
            a_sheet_exists = "-"
            a_data_exists = "-"

        # B_sampleシートの確認
        if "B_sample" in wb_in.sheetnames:
            b_sheet_exists = "v"
            b_data_exists = "v" if wb_in["B_sample"]["C3"].value else "-"
        else:
            b_sheet_exists = "-"
            b_data_exists = "-"

        # 結果をワークシートに追加
        ws.append([file_name, a_sheet_exists, a_data_exists, b_sheet_exists, b_data_exists])

        wb_in.close()

# 列幅の自動調整
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# ファイル名を現在の日付・時間に基づいて生成
output_file_name = "Check_Sheet_" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"

# ファイルを出力ディレクトリに保存
output_file_path = os.path.join(output_dir, output_file_name)
wb.save(output_file_path)